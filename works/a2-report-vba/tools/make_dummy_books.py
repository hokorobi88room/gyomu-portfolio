"""ダミー月次報告ブックの生成スクリプト。

20支店分の報告ブックを sample-data/reports/ へ生成する。
実際の現場を模して、支店ごとに様式を意図的に揺らがせる:
- ヘッダー行の位置(1〜4行目)
- 列の並び順
- ヘッダーの別名(売上/金額/売上高、商品/品目/商品名)
- 小計行・メモ行の混入
- 1冊はヘッダー不備(取込レポートの「警告」動作確認用)
"""
from __future__ import annotations

import random
from datetime import date
from pathlib import Path

from openpyxl import Workbook

OUT_DIR = Path(__file__).resolve().parent.parent / "sample-data" / "reports"

BRANCHES = [
    "札幌", "仙台", "新潟", "大宮", "千葉", "東京", "横浜", "静岡", "名古屋", "金沢",
    "京都", "大阪", "神戸", "岡山", "広島", "高松", "福岡", "熊本", "鹿児島", "那覇",
]
STAFF = ["佐藤", "鈴木", "高橋", "田中", "伊藤", "渡辺", "山本", "中村"]
ITEMS = ["スタンダードプラン", "プレミアムプラン", "保守サポート", "初期構築", "オプション追加"]

DATE_ALIASES = ["日付", "売上日", "取引日"]
STAFF_ALIASES = ["担当者", "担当", "担当者名"]
ITEM_ALIASES = ["商品", "品目", "商品名"]
SALES_ALIASES = ["売上", "金額", "売上高", "売上(円)"]


def make_book(branch: str, index: int, rng: random.Random) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "月次報告"

    header_row = rng.choice([1, 2, 3, 4])
    for r in range(1, header_row):
        ws.cell(row=r, column=1, value=f"{branch}支店 2026年6月 月次報告" if r == 1 else "")

    headers = [
        (rng.choice(DATE_ALIASES), "date"),
        (rng.choice(STAFF_ALIASES), "staff"),
        (rng.choice(ITEM_ALIASES), "item"),
        (rng.choice(SALES_ALIASES), "sales"),
    ]
    rng.shuffle(headers)

    broken = index == 19  # 最後の1冊はヘッダー不備(警告動作の確認用)
    for c, (label, kind) in enumerate(headers, start=1):
        if broken and kind == "sales":
            label = "備考"  # 売上列が存在しない報告書
        ws.cell(row=header_row, column=c, value=label)

    col_of = {kind: c for c, (_, kind) in enumerate(headers, start=1)}
    r = header_row + 1
    for _ in range(rng.randint(15, 40)):
        d = date(2026, rng.choice([4, 5, 6]), rng.randint(1, 28))
        ws.cell(row=r, column=col_of["date"], value=d)
        ws.cell(row=r, column=col_of["staff"], value=rng.choice(STAFF))
        ws.cell(row=r, column=col_of["item"], value=rng.choice(ITEMS))
        ws.cell(row=r, column=col_of["sales"], value=rng.randint(2, 60) * 5000)
        r += 1
        if rng.random() < 0.08:  # 小計行の混入
            ws.cell(row=r, column=col_of["item"], value="―― 小計 ――")
            r += 1

    ws.cell(row=r + 1, column=1, value="※ 系列システムから出力後、手修正あり")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_DIR / f"月次報告_{branch}支店_202606.xlsx")


def main() -> None:
    rng = random.Random(42)  # 再現可能な生成
    for i, branch in enumerate(BRANCHES):
        make_book(branch, i, rng)
    print(f"generated {len(BRANCHES)} books -> {OUT_DIR}")


if __name__ == "__main__":
    main()
