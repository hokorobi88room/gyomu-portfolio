"""WebWatch のテスト(ネットワーク不使用: 抽出・差分・Excel・設定・失敗系)。"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from webwatch.config import Config, ConfigError, Target, load_config  # noqa: E402
from webwatch.differ import Diff, SnapshotStore, diff_rows  # noqa: E402
from webwatch.excel_out import write_report  # noqa: E402
from webwatch.extractor import extract  # noqa: E402

FIXTURE_HTML = """
<html><body>
  <ul class="release-list">
    <li><span class="date">2026-07-01</span><a class="title">6月統計を公開</a><span class="category">月次</span></li>
    <li><span class="date">2026-06-24</span><a class="title">年報を更新</a><span class="category">年次</span></li>
    <li><span class="date"></span><a class="title"></a><span class="category"></span></li>
  </ul>
</body></html>
"""

TARGET = Target(
    name="統計新着",
    url="https://example.com/stats",
    item_selector="ul.release-list li",
    fields={"日付": ".date", "タイトル": "a.title", "区分": ".category"},
)


# ---------- 抽出 ----------

def test_extract_rows_and_skip_empty_blocks() -> None:
    rows = extract(FIXTURE_HTML, TARGET)
    assert len(rows) == 2  # 全フィールド空のブロックは除外
    assert rows[0] == {"日付": "2026-07-01", "タイトル": "6月統計を公開", "区分": "月次"}


def test_extract_missing_field_becomes_empty_string() -> None:
    target = Target(
        name="t", url="https://example.com", item_selector="li",
        fields={"日付": ".date", "存在しない": ".nope"},
    )
    rows = extract(FIXTURE_HTML, target)
    assert all(r["存在しない"] == "" for r in rows)
    assert rows[0]["日付"] == "2026-07-01"


# ---------- 差分 ----------

def test_diff_added_and_removed() -> None:
    prev = [{"a": "1"}, {"a": "2"}]
    curr = [{"a": "2"}, {"a": "3"}]
    d = diff_rows(prev, curr)
    assert d.added == [{"a": "3"}]
    assert d.removed == [{"a": "1"}]
    assert d.has_changes


def test_snapshot_roundtrip_and_broken_file(tmp_path: Path) -> None:
    store = SnapshotStore(tmp_path)
    store.save("対象/名", [{"a": "1"}])            # 記号入りの名前も安全なファイル名になる
    assert store.load("対象/名") == [{"a": "1"}]
    # 壊れたスナップショットは初回扱い(例外を投げない)
    next(tmp_path.glob("*.json")).write_text("{{broken", encoding="utf-8")
    assert store.load("対象/名") == []


# ---------- Excel出力 ----------

def test_write_report_marks_new_rows(tmp_path: Path) -> None:
    results = {"統計新着": [{"日付": "2026-07-01", "タイトル": "6月統計"}]}
    diffs = {"統計新着": Diff(added=[{"日付": "2026-07-01", "タイトル": "6月統計"}])}
    path = write_report(results, diffs, tmp_path, now=datetime(2026, 7, 3, 12, 0))
    wb = load_workbook(path)
    assert "差分サマリ" in wb.sheetnames and "統計新着" in wb.sheetnames
    ws = wb["統計新着"]
    assert ws.cell(row=2, column=3).value == "NEW"
    summary = wb["差分サマリ"]
    assert summary.cell(row=2, column=3).value == 1  # 新規1件


def test_write_report_empty_target(tmp_path: Path) -> None:
    path = write_report({"空対象": []}, {"空対象": Diff()}, tmp_path)
    ws = load_workbook(path)["空対象"]
    assert ws.cell(row=1, column=1).value == "取得0件"


# ---------- 設定(異常系) ----------

def test_load_config_ok(tmp_path: Path) -> None:
    p = tmp_path / "c.yaml"
    p.write_text(
        "targets:\n"
        "  - name: t\n    url: https://example.com\n    item_selector: li\n"
        "    fields:\n      日付: .date\n",
        encoding="utf-8",
    )
    cfg = load_config(p)
    assert isinstance(cfg, Config)
    assert cfg.targets[0].name == "t"


@pytest.mark.parametrize(
    ("body", "match"),
    [
        ("targets: []", "1件以上"),
        ("targets:\n  - name: t\n    url: ftp://x\n    item_selector: li\n    fields: {a: b}", "http"),
        ("targets:\n  - name: t\n    url: https://x\n    item_selector: li\n    fields: {a: b}\ninterval_seconds: 0.1", "1秒以上"),
        ("targets:\n  - url: https://x\n    item_selector: li\n    fields: {a: b}", "name"),
    ],
)
def test_load_config_errors(tmp_path: Path, body: str, match: str) -> None:
    p = tmp_path / "c.yaml"
    p.write_text(body, encoding="utf-8")
    with pytest.raises(ConfigError, match=match):
        load_config(p)


def test_load_config_missing_file(tmp_path: Path) -> None:
    with pytest.raises(ConfigError, match="ありません"):
        load_config(tmp_path / "nai.yaml")
