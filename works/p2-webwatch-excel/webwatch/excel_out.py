"""体裁の整ったExcelレポート出力(openpyxl)。"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .differ import Diff

HEADER_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ADDED_FILL = PatternFill("solid", fgColor="D1FAE5")
REMOVED_FILL = PatternFill("solid", fgColor="FEE2E2")


def write_report(
    results: dict[str, list[dict[str, str]]],
    diffs: dict[str, Diff],
    out_dir: str | Path,
    now: datetime | None = None,
) -> Path:
    """ターゲットごとに1シート+差分サマリシートのブックを出力する。"""
    now = now or datetime.now()
    wb = Workbook()

    # --- サマリシート ---
    ws = wb.active
    ws.title = "差分サマリ"
    _headers(ws, ["対象", "取得件数", "新規", "消滅", "状態"])
    r = 2
    for name, rows in results.items():
        d = diffs.get(name, Diff())
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=len(rows))
        ws.cell(row=r, column=3, value=len(d.added))
        ws.cell(row=r, column=4, value=len(d.removed))
        status = "変化あり" if d.has_changes else "変化なし"
        c = ws.cell(row=r, column=5, value=status)
        if d.has_changes:
            c.font = Font(bold=True, color="B45309")
        r += 1
    ws.cell(row=r + 1, column=1, value=f"取得日時: {now:%Y/%m/%d %H:%M}")
    _autofit(ws)

    # --- ターゲット別シート ---
    for name, rows in results.items():
        wst = wb.create_sheet(title=name[:31])  # Excelのシート名は31文字まで
        d = diffs.get(name, Diff())
        added_keys = {tuple(sorted(x.items())) for x in d.added}

        if not rows:
            wst.cell(row=1, column=1, value="取得0件")
            continue
        columns = list(rows[0].keys())
        _headers(wst, [*columns, "状態"])
        for i, row in enumerate(rows, start=2):
            for j, col in enumerate(columns, start=1):
                wst.cell(row=i, column=j, value=row.get(col, ""))
            is_new = tuple(sorted(row.items())) in added_keys
            wst.cell(row=i, column=len(columns) + 1, value="NEW" if is_new else "")
            if is_new:
                for j in range(1, len(columns) + 2):
                    wst.cell(row=i, column=j).fill = ADDED_FILL
        # 消滅行を末尾に赤で記録(黙って消さない)
        base = len(rows) + 3
        if d.removed:
            wst.cell(row=base - 1, column=1, value="― 前回から消滅した行 ―").font = Font(bold=True)
            for i, row in enumerate(d.removed, start=base):
                for j, col in enumerate(columns, start=1):
                    c = wst.cell(row=i, column=j, value=row.get(col, ""))
                    c.fill = REMOVED_FILL
        wst.auto_filter.ref = f"A1:{get_column_letter(len(columns) + 1)}{len(rows) + 1}"
        _autofit(wst)

    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    path = out / f"webwatch_{now:%Y%m%d_%H%M}.xlsx"
    wb.save(path)
    return path


def _headers(ws: Worksheet, labels: list[str]) -> None:
    for j, label in enumerate(labels, start=1):
        c = ws.cell(row=1, column=j, value=label)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autofit(ws: Worksheet) -> None:
    for column_cells in ws.columns:
        width = max(
            (len(str(c.value)) * 1.8 if any(ord(ch) > 0x7F for ch in str(c.value)) else len(str(c.value)))
            if c.value is not None else 0
            for c in column_cells
        )
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width + 2, 10), 60)
